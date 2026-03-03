import os
import sys
import time
import json
import html
import traceback
import re
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Tuple, Optional

import requests
import feedparser
from openpyxl import load_workbook

import numpy as np
import pandas as pd
import yfinance as yf


VERSION = "0.5.0"
TELEGRAM_API = "https://api.telegram.org"

ETF_MAP_PATH = Path("config/etf_map.json")
STATE_PATH = Path("data/state.json")

# Free daily OHLCV (best-effort)
STOOQ_BASE = "https://stooq.com/q/d/l/"
# SPDR (SSGA) daily holdings XLSX (free)
SSGA_HOLDINGS_TMPL = "https://www.ssga.com/library-content/products/fund-data/etfs/us/holdings-daily-us-en-{}.xlsx"

HANGUL_RE = re.compile(r"[가-힣]")


# -----------------------
# Helpers
# -----------------------
def now_kst() -> datetime:
    return datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=9)))


def env(name: str, default: str = "") -> str:
    v = os.getenv(name, default)
    return v.strip() if isinstance(v, str) else default


def env_int(name: str, default: int) -> int:
    try:
        return int(float(env(name, str(default))))
    except Exception:
        return default


def env_float(name: str, default: float) -> float:
    try:
        return float(env(name, str(default)))
    except Exception:
        return default


def get_lang() -> str:
    lang = (env("BRIEF_LANG", "ko") or "ko").lower()
    return "en" if lang.startswith("en") else "ko"


def tr(key: str) -> str:
    ko = {
        "title": "데일리 브리핑",
        "status_ok": "정상",
        "status_err": "오류",
        "etf_loaded": "ETF 지도 로드",

        "s2": "S2. ETF 지도",
        "benchmark": "벤치마크",
        "source": "소스",
        "sector_a": "섹터 A (지속 강세)",
        "sector_b": "섹터 B (회복 시도)",
        "country_a": "국가/지역 A (지속 강세)",
        "country_b": "국가/지역 B (회복 시도)",

        "s3a": "S3-A. 진입 임박 후보(0~4)",
        "s3b": "S3-B. 바닥 형성 후보(승급 대기)",
        "s3_none": "• 오늘 후보 없음 (조건 미충족/데이터 부족)",
        "s3_note": "※ 후보는 'ENTRY(100/50) 맹신'이 아니라 '메인지표 상태(ST DOWN + arm 준비도)' 기반. 최종 진입은 TradingView 4H 메인지표로 확인.",

        "s1": "S1. 오늘 체크(제목 기반 요약)",
        "s1_note": "※ 본문 요약이 아니라 제목 키워드로 '리스크/이벤트'만 압축. (원문 클릭은 선택)",
        "no_news": "• (시장 관련 헤드라인 없음)",
        "raw_headlines": "원문 헤드라인",

        "s4_log": "S4. 로그",
        "s4_err": "S4. 에러",
        "state_missing": "• state.json이 없어 새로 생성(정상)",
    }
    en = {"title": "Daily Briefing", "status_ok": "OK", "status_err": "ERROR"}
    table = ko if get_lang() == "ko" else en
    return table.get(key, key)


def truncate_telegram(text: str, limit: int = 3900) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n…(truncated)"


# -----------------------
# State
# -----------------------
def load_state() -> Tuple[dict, List[str]]:
    logs = []
    default_state = {"last_run_kst": "", "seen_links": [], "candidate_history": []}
    try:
        if not STATE_PATH.exists():
            STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
            STATE_PATH.write_text(json.dumps(default_state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
            logs.append(tr("state_missing"))
            return default_state, logs
        data = json.loads(STATE_PATH.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return default_state, logs
        data.setdefault("last_run_kst", "")
        data.setdefault("seen_links", [])
        data.setdefault("candidate_history", [])
        return data, logs
    except Exception:
        logs.append("• state.json 로드 실패(무시하고 진행)")
        return default_state, logs


def save_state(state: dict) -> None:
    try:
        STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
        STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    except Exception:
        print("[WARN] Failed to save state.json", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)


def prune_state(state: dict) -> None:
    max_links = env_int("STATE_MAX_LINKS", 1500)
    seen = state.get("seen_links", []) or []
    if len(seen) > max_links:
        state["seen_links"] = seen[-max_links:]


def mark_links_seen(state: dict, links: List[str]) -> None:
    seen = state.get("seen_links", []) or []
    for l in links:
        if l and l not in seen:
            seen.append(l)
    state["seen_links"] = seen
    prune_state(state)


def add_candidate_history(state: dict, items: List[dict]) -> None:
    hist = state.get("candidate_history", []) or []
    today = now_kst().date()
    cutoff = today - timedelta(days=60)

    kept = []
    for h in hist:
        try:
            d = datetime.strptime(h.get("date_kst", ""), "%Y-%m-%d").date()
            if d >= cutoff:
                kept.append(h)
        except Exception:
            continue

    kept.extend(items)
    state["candidate_history"] = kept


def was_recently_recommended(state: dict, ticker: str, days: int) -> bool:
    hist = state.get("candidate_history", []) or []
    today = now_kst().date()
    for h in reversed(hist):
        if h.get("ticker") == ticker:
            try:
                d = datetime.strptime(h.get("date_kst", ""), "%Y-%m-%d").date()
                return (today - d).days <= days
            except Exception:
                return True
    return False


# -----------------------
# Config / ETF map / Holdings
# -----------------------
def load_etf_map() -> dict:
    fallback = {"benchmark": ["ACWI"], "sectors_11": [], "countries_regions_16": {"developed": [], "emerging": []}}
    try:
        with ETF_MAP_PATH.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        print("[WARN] Failed to load config/etf_map.json", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        return fallback


def fetch_spdr_holdings(etf_ticker: str, timeout: int = 12) -> List[dict]:
    t = etf_ticker.strip().lower()
    url = SSGA_HOLDINGS_TMPL.format(t)
    headers = {"User-Agent": env("USER_AGENT", "RoyBriefbot/0.5.0 (+GitHub Actions)")}
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()

    wb = load_workbook(filename=BytesIO(resp.content), read_only=True, data_only=True)
    ws = wb.active

    header_row = None
    col_ticker = None
    col_name = None
    col_weight = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=80, values_only=True), start=1):
        if not row:
            continue
        norm = [str(c).strip().lower() if c is not None else "" for c in row]
        if any(("ticker" in x) or (x == "symbol") for x in norm) and any("weight" in x for x in norm):
            header_row = i
            for j, x in enumerate(norm):
                if col_ticker is None and ("ticker" in x or x == "symbol"):
                    col_ticker = j
                if col_name is None and ("name" == x or "security name" in x or x.endswith("name")):
                    col_name = j
                if col_weight is None and ("weight" in x):
                    col_weight = j
            break

    if header_row is None or col_ticker is None or col_weight is None:
        return []

    out: List[dict] = []
    blank = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 5000, values_only=True):
        if not row:
            blank += 1
            if blank >= 20:
                break
            continue

        raw_t = row[col_ticker] if col_ticker < len(row) else None
        raw_w = row[col_weight] if col_weight < len(row) else None
        raw_n = row[col_name] if (col_name is not None and col_name < len(row)) else None

        if raw_t is None or str(raw_t).strip() == "":
            blank += 1
            if blank >= 20:
                break
            continue

        blank = 0
        ticker = str(raw_t).strip().upper()
        if len(ticker) > 10 or not ticker[0].isalpha():
            continue

        try:
            weight = float(raw_w)
        except Exception:
            try:
                weight = float(str(raw_w).replace("%", "").strip())
            except Exception:
                continue

        name = str(raw_n).strip() if raw_n is not None else ""
        out.append({"ticker": ticker, "name": name, "weight": weight})

    out.sort(key=lambda x: x.get("weight", 0.0), reverse=True)
    return out


# -----------------------
# News (RSS)
# -----------------------
def load_rss_urls() -> List[str]:
    raw = env("RSS_URLS")
    if raw:
        return [u.strip() for u in raw.split(",") if u.strip()]
    return [
        "https://www.hankyung.com/feed/finance",
        "https://www.hankyung.com/feed/economy",
        "https://www.yonhapnewseconomytv.com/rss/allArticle.xml",
        "https://news.google.com/rss/headlines/section/topic/BUSINESS?hl=ko&gl=KR&ceid=KR:ko",
    ]


def _keyword_list() -> List[str]:
    custom = env("NEWS_KEYWORDS")
    if custom:
        return [k.strip() for k in custom.split(",") if k.strip()]
    return [
        "FOMC", "연준", "금리", "CPI", "PCE", "고용", "실업", "국채", "채권", "달러", "환율",
        "실적", "어닝", "가이던스", "반도체", "AI", "유가", "OPEC",
        "코스피", "코스닥", "나스닥", "S&P", "VIX", "리밸런싱", "지정학", "전쟁", "제재",
    ]


def is_korean_title(title: str) -> bool:
    return bool(HANGUL_RE.search(title or ""))


def fetch_headlines(rss_urls: List[str], top_n: int = 16, timeout: int = 12, state: Optional[dict] = None) -> Tuple[List[dict], List[str]]:
    items = []
    logs = []
    per_feed = []

    for url in rss_urls:
        try:
            headers = {"User-Agent": env("USER_AGENT", "RoyBriefbot/0.5.0 (+GitHub Actions)")}
            resp = requests.get(url, headers=headers, timeout=timeout)
            resp.raise_for_status()
            feed = feedparser.parse(resp.text)
            source = (feed.feed.get("title") or url).strip()

            count = 0
            for e in feed.entries[: max(top_n, 80)]:
                title = (e.get("title") or "").strip()
                link = (e.get("link") or "").strip()
                if not title:
                    continue

                ts = 0
                if "published_parsed" in e and e.published_parsed:
                    ts = int(time.mktime(e.published_parsed))
                elif "updated_parsed" in e and e.updated_parsed:
                    ts = int(time.mktime(e.updated_parsed))

                items.append({"source": source, "title": title, "link": link, "_ts": ts})
                count += 1

            per_feed.append((source, count))
        except Exception:
            logs.append(f"RSS 실패: {url}")
            per_feed.append((url, 0))

    ok_cnt = sum(1 for _, c in per_feed if c > 0)
    logs.append(f"RSS 상태: {ok_cnt}/{len(per_feed)} 피드 수집")

    items.sort(key=lambda x: x.get("_ts", 0), reverse=True)
    seen_title = set()
    dedup = []
    for it in items:
        key = it["title"].lower()
        if key in seen_title:
            continue
        seen_title.add(key)
        it.pop("_ts", None)
        dedup.append(it)

    if env("NEWS_FILTER", "1") != "0":
        kws = [k.lower() for k in _keyword_list()]
        filtered = []
        for it in dedup:
            t = it["title"].lower()
            if any(k in t for k in kws):
                filtered.append(it)
        dedup = filtered if len(filtered) >= 3 else dedup

    if get_lang() == "ko" and env("NEWS_REQUIRE_KOREAN", "1") != "0":
        ko_only = [it for it in dedup if is_korean_title(it.get("title", ""))]
        if len(ko_only) >= 2:
            dedup = ko_only
        else:
            logs.append("한글 뉴스 부족: 일부 영문이 섞일 수 있음")

    if state is not None and env("STATE_DEDUPE_NEWS", "1") != "0":
        seen_links = set(state.get("seen_links", []) or [])
        fresh = [it for it in dedup if it.get("link") and it["link"] not in seen_links]
        if len(fresh) >= 3:
            return fresh[:top_n], logs
        return (fresh + dedup)[:top_n], logs

    return dedup[:top_n], logs


def news_score(title: str) -> Tuple[int, List[str], str]:
    t = (title or "").lower()
    tags: List[str] = []
    score = 0

    def hit(words: List[str], pts: int, tag: str):
        nonlocal score
        if any(w.lower() in t for w in words):
            score += pts
            tags.append(tag)

    hit(["fomc", "fed", "연준"], 6, "🟧통화정책")
    hit(["cpi", "pce", "inflation", "물가"], 6, "🟧물가지표")
    hit(["jobs", "employment", "실업", "고용"], 5, "🟨고용지표")
    hit(["yield", "treasury", "국채", "채권"], 5, "🟨금리/채권")
    hit(["usd", "dollar", "환율", "달러"], 4, "🟨환율")
    hit(["earnings", "guidance", "실적", "어닝", "가이던스"], 5, "🟦실적")
    hit(["vix", "volatility", "변동성"], 5, "🟥변동성")
    hit(["oil", "opec", "유가"], 4, "🟩원자재")
    hit(["war", "전쟁", "제재", "지정학"], 5, "🟥지정학")
    hit(["rebalanc", "리밸런싱", "월마감"], 4, "🟪수급")
    hit(["semiconductor", "chip", "반도체"], 4, "🟦섹터")
    hit(["ai", "인공지능"], 3, "🟦테마")

    if "🟥" in "".join(tags):
        note = "변동성/리스크 이벤트 가능. 신규 진입 보수."
    elif "🟧통화정책" in tags or "🟧물가지표" in tags or "🟨고용지표" in tags:
        note = "매크로 변동 구간. 사이즈/무효화 기준 엄격."
    elif "🟦실적" in tags:
        note = "실적 갭 리스크. 익일 시초가 변동 주의."
    elif "🟪수급" in tags:
        note = "월말/리밸런싱 수급 왜곡 가능."
    else:
        note = "시장 영향 가능. 필요시 원문 확인."

    tags = sorted(set(tags))
    return score, tags, note


def build_news_digest(headlines: List[dict]) -> Tuple[List[str], List[str]]:
    top_k = env_int("NEWS_DIGEST_TOP", 3)
    show_raw = env("SHOW_RAW_HEADLINES", "0") != "0"

    scored = []
    for h in headlines:
        s, tags, note = news_score(h.get("title", ""))
        scored.append((s, tags, note, h))

    scored.sort(key=lambda x: x[0], reverse=True)
    picked = [x for x in scored if x[0] > 0][:top_k]
    if not picked:
        picked = scored[:min(top_k, len(scored))]

    lines: List[str] = []
    lines.append(f"<b>{html.escape(tr('s1'))}</b>")
    lines.append(html.escape(tr("s1_note")))
    if not picked:
        lines.append(html.escape(tr("no_news")))
        return lines, []

    used_links: List[str] = []
    for i, (s, tags, note, h) in enumerate(picked, 1):
        title = h.get("title", "")
        link = h.get("link", "")
        src = h.get("source", "")
        tag_str = " ".join(tags) if tags else "🟦시장"
        if link:
            lines.append(f"{i}) {tag_str} <a href=\"{html.escape(link)}\">{html.escape(title)}</a> <i>({html.escape(src)})</i>")
            used_links.append(link)
        else:
            lines.append(f"{i}) {tag_str} {html.escape(title)} <i>({html.escape(src)})</i>")
        lines.append(f"   • {html.escape(note)}")

    if show_raw:
        lines.append("")
        lines.append(f"<b>{html.escape(tr('raw_headlines'))}</b>")
        for i, h in enumerate(headlines[:8], 1):
            src = html.escape(h.get("source", ""))
            title = html.escape(h.get("title", ""))
            link = h.get("link", "")
            if link:
                lines.append(f'{i}) <a href="{html.escape(link)}">{title}</a> <i>({src})</i>')
            else:
                lines.append(f"{i}) {title} <i>({src})</i>")

    return lines, used_links


# -----------------------
# S2: ETF map (daily RS)
# -----------------------
def _stooq_symbol(ticker: str) -> str:
    t = ticker.strip().lower()
    if t.endswith(".us") or t.endswith(".uk") or t.endswith(".jp") or t.endswith(".de") or t.endswith(".pl"):
        return t
    return f"{t}.us"


def _fetch_stooq_daily_ohlcv(ticker: str, timeout: int = 12) -> List[dict]:
    sym = _stooq_symbol(ticker)
    url = f"{STOOQ_BASE}?s={sym}&i=d"
    headers = {"User-Agent": env("USER_AGENT", "RoyBriefbot/0.5.0 (+GitHub Actions)")}
    resp = requests.get(url, headers=headers, timeout=timeout)
    resp.raise_for_status()

    text = resp.text.strip()
    if not text or "Date,Open,High,Low,Close,Volume" not in text:
        return []

    lines = text.splitlines()
    if len(lines) < 3:
        return []

    out: List[dict] = []
    for row in lines[1:]:
        cols = row.split(",")
        if len(cols) < 6:
            continue
        try:
            out.append({"date": cols[0].strip(), "open": float(cols[1]), "high": float(cols[2]), "low": float(cols[3]), "close": float(cols[4]), "volume": float(cols[5])})
        except Exception:
            continue
    return out


def _closes(series: List[dict]) -> List[float]:
    return [x["close"] for x in series]


def _pct_change(closes: List[float], lookback: int) -> Optional[float]:
    if len(closes) < lookback + 1:
        return None
    last = closes[-1]
    prev = closes[-(lookback + 1)]
    if prev == 0:
        return None
    return (last / prev - 1.0) * 100.0


def build_etf_rankings(tickers: List[str], benchmark: str, timeout: int = 12) -> Tuple[List[dict], List[str]]:
    failed: List[str] = []
    cache: Dict[str, List[dict]] = {}

    def get_series(t: str) -> List[dict]:
        if t in cache:
            return cache[t]
        s = _fetch_stooq_daily_ohlcv(t, timeout=timeout)
        cache[t] = s
        return s

    bench_series = get_series(benchmark)
    bench_closes = _closes(bench_series) if bench_series else []
    if not bench_closes:
        failed.append(benchmark)

    horizons = {"1w": 5, "1m": 21, "3m": 63, "6m": 126}
    bench_ret = {k: _pct_change(bench_closes, v) for k, v in horizons.items()} if bench_closes else {}

    rows: List[dict] = []
    for t in tickers:
        try:
            series = get_series(t)
            closes = _closes(series) if series else []
            if not closes:
                failed.append(t)
                continue
            ret = {k: _pct_change(closes, v) for k, v in horizons.items()}
            rs = {}
            for k in horizons.keys():
                if bench_closes and ret.get(k) is not None and bench_ret.get(k) is not None:
                    rs[k] = ret[k] - bench_ret[k]
                else:
                    rs[k] = None
            rows.append({"ticker": t, "rs": rs})
        except Exception:
            failed.append(t)

    return rows, failed


def pick_top_tracks(rows: List[dict], top_n: int = 3) -> Tuple[List[dict], List[dict]]:
    def score_a(r: dict) -> float:
        rs6 = r["rs"].get("6m")
        rs3 = r["rs"].get("3m")
        if rs6 is None or rs3 is None:
            return -1e9
        return 0.6 * rs6 + 0.4 * rs3

    def score_b(r: dict) -> float:
        rs1w = r["rs"].get("1w")
        rs1m = r["rs"].get("1m")
        rs3m = r["rs"].get("3m")
        if rs1w is None or rs1m is None or rs3m is None:
            return -1e9
        if rs1w <= 0:
            return -1e9
        bonus = 0.5 if rs3m < 0 else 0.0
        return (rs1w - rs1m) + bonus

    a_sorted = sorted(rows, key=score_a, reverse=True)
    b_sorted = sorted(rows, key=score_b, reverse=True)
    top_a = [r for r in a_sorted if score_a(r) > -1e8][:top_n]
    top_b = [r for r in b_sorted if score_b(r) > -1e8][:top_n]
    return top_a, top_b


def compute_s2_picks(etf_map: dict) -> Tuple[List[str], List[str], List[str], List[str], List[str]]:
    warnings: List[str] = []
    benchmark = (etf_map.get("benchmark", []) or ["ACWI"])[0]
    sectors = etf_map.get("sectors_11", []) or []
    cr = etf_map.get("countries_regions_16", {}) or {}
    countries = (cr.get("developed", []) or []) + (cr.get("emerging", []) or [])
    timeout = env_int("DATA_TIMEOUT", 12)

    sector_a: List[str] = []
    sector_b: List[str] = []
    country_a: List[str] = []
    country_b: List[str] = []

    if sectors:
        rows, failed = build_etf_rankings(sectors, benchmark=benchmark, timeout=timeout)
        if failed:
            warnings.append("S2 sectors data missing: " + ", ".join(sorted(set(failed))[:8]))
        top_a, top_b = pick_top_tracks(rows, top_n=3)
        sector_a = [r["ticker"] for r in top_a]
        sector_b = [r["ticker"] for r in top_b]

    if countries:
        rows, failed = build_etf_rankings(countries, benchmark=benchmark, timeout=timeout)
        if failed:
            warnings.append("S2 countries data missing: " + ", ".join(sorted(set(failed))[:8]))
        top_a, top_b = pick_top_tracks(rows, top_n=3)
        country_a = [r["ticker"] for r in top_a]
        country_b = [r["ticker"] for r in top_b]

    return sector_a, sector_b, country_a, country_b, warnings


def build_s2_section(etf_map: dict, sector_a: List[str], sector_b: List[str], country_a: List[str], country_b: List[str]) -> List[str]:
    benchmark = (etf_map.get("benchmark", []) or ["ACWI"])[0]
    lines: List[str] = []
    lines.append(f"<b>{tr('s2')}</b>")
    lines.append(f"<i>{html.escape(tr('benchmark'))}</i>: {html.escape(benchmark)}  <i>{html.escape(tr('source'))}</i>: Stooq (daily)")
    lines.append("")
    lines.append(f"<b>{tr('sector_a')}</b>")
    lines.extend([f"• {html.escape(t)}" for t in sector_a] or ["• (없음)"])
    lines.append(f"<b>{tr('sector_b')}</b>")
    lines.extend([f"• {html.escape(t)}" for t in sector_b] or ["• (없음)"])
    lines.append("")
    lines.append(f"<b>{tr('country_a')}</b>")
    lines.extend([f"• {html.escape(t)}" for t in country_a] or ["• (없음)"])
    lines.append(f"<b>{tr('country_b')}</b>")
    lines.extend([f"• {html.escape(t)}" for t in country_b] or ["• (없음)"])
    return lines


# -----------------------
# 4H Main-indicator replica (approx, Pine parity for WT/MFI/BB/ST)
# -----------------------
def _ema(series: np.ndarray, n: int) -> np.ndarray:
    # Exponential moving average (alpha=2/(n+1))
    if n <= 1:
        return series.copy()
    alpha = 2.0 / (n + 1.0)
    out = np.empty_like(series, dtype=float)
    out[:] = np.nan
    acc = series[0]
    out[0] = acc
    for i in range(1, len(series)):
        acc = alpha * series[i] + (1 - alpha) * acc
        out[i] = acc
    return out


def _sma(series: np.ndarray, n: int) -> np.ndarray:
    if n <= 1:
        return series.copy()
    out = pd.Series(series).rolling(n, min_periods=n).mean().to_numpy(dtype=float)
    return out


def _stdev(series: np.ndarray, n: int) -> np.ndarray:
    out = pd.Series(series).rolling(n, min_periods=n).std(ddof=0).to_numpy(dtype=float)
    return out


def _rma(series: np.ndarray, n: int) -> np.ndarray:
    # Wilder's RMA, used by ta.atr
    if n <= 1:
        return series.copy()
    out = np.empty_like(series, dtype=float)
    out[:] = np.nan
    out[0] = series[0]
    alpha = 1.0 / n
    for i in range(1, len(series)):
        out[i] = out[i-1] + alpha * (series[i] - out[i-1])
    return out


def _atr(high: np.ndarray, low: np.ndarray, close: np.ndarray, n: int) -> np.ndarray:
    tr = np.empty_like(close, dtype=float)
    tr[:] = np.nan
    tr[0] = high[0] - low[0]
    for i in range(1, len(close)):
        tr[i] = max(high[i] - low[i], abs(high[i] - close[i-1]), abs(low[i] - close[i-1]))
    return _rma(tr, n)


def _mfi(high: np.ndarray, low: np.ndarray, close: np.ndarray, vol: np.ndarray, n: int) -> np.ndarray:
    tp = (high + low + close) / 3.0
    mf = tp * vol
    pos = np.zeros_like(tp, dtype=float)
    neg = np.zeros_like(tp, dtype=float)
    for i in range(1, len(tp)):
        if tp[i] > tp[i-1]:
            pos[i] = mf[i]
        elif tp[i] < tp[i-1]:
            neg[i] = mf[i]
    pos_sum = pd.Series(pos).rolling(n, min_periods=n).sum().to_numpy(dtype=float)
    neg_sum = pd.Series(neg).rolling(n, min_periods=n).sum().to_numpy(dtype=float)
    mfi = np.full_like(tp, np.nan, dtype=float)
    for i in range(len(tp)):
        if np.isnan(pos_sum[i]) or np.isnan(neg_sum[i]):
            continue
        if neg_sum[i] == 0:
            mfi[i] = 100.0
        else:
            r = pos_sum[i] / neg_sum[i]
            mfi[i] = 100.0 - (100.0 / (1.0 + r))
    return mfi


def yf_fetch_1h(ticker: str, period: str, interval: str) -> Optional[pd.DataFrame]:
    try:
        df = yf.download(
            tickers=ticker,
            period=period,
            interval=interval,
            auto_adjust=False,
            progress=False,
            threads=False,
            prepost=False
        )
        if df is None or df.empty:
            return None
        # Normalize columns
        cols = [c.lower() for c in df.columns]
        df.columns = cols
        for need in ["open", "high", "low", "close", "volume"]:
            if need not in df.columns:
                return None
        df = df[["open", "high", "low", "close", "volume"]].dropna()
        if len(df) < 100:
            return None
        return df
    except Exception:
        return None


def resample_4h(df_1h: pd.DataFrame) -> Optional[pd.DataFrame]:
    try:
        df = df_1h.copy()
        df.index = pd.to_datetime(df.index)
        # 4H bars by time buckets
        agg = {
            "open": "first",
            "high": "max",
            "low": "min",
            "close": "last",
            "volume": "sum",
        }
        df4 = df.resample("4H").agg(agg).dropna()
        # basic sanity
        if len(df4) < 120:
            return None
        return df4
    except Exception:
        return None


def compute_main_setup(df4: pd.DataFrame) -> Optional[dict]:
    """
    Replicates the MAIN indicator's 'state/arm' logic (WT/MFI/BB/ST) on 4H bars.
    Outputs a compact snapshot for scanning (NOT an execution signal).
    """
    # Params from Pine GLOBAL preset (env override)
    wt_n1 = env_int("WT_N1", 24)
    wt_n2 = env_int("WT_N2", 41)
    wt_n3 = env_int("WT_N3", 5)
    wt_ob = env_float("WT_OB", 50.5)
    wt_os = env_float("WT_OS", -41.0)

    mfi_len = env_int("MFI_LEN", 11)
    mfi_smooth = env_int("MFI_SMOOTH", 6)
    mfi_linger = env_int("MFI_LINGER", 1)
    mfi_ob = env_float("MFI_OB", 68.0)
    mfi_os = env_float("MFI_OS", 32.0)

    bb_len = env_int("BB_LEN", 20)
    bb_mult = env_float("BB_MULT", 2.2070048992556504)
    bb_rein = max(env_int("BB_REIN", 2), 1)
    bb_reentry = env_int("BB_REENTRY_BARS", 6)

    st_atr_len = env_int("ST_ATR_LEN", 8)
    st_factor = env_float("ST_FACTOR", 4.325)
    st_confirm = max(env_int("ST_CONFIRM", 1), 1)
    st_hys_atr = env_float("ST_HYS_ATR", 0.275)

    ttl = env_int("TTL_CLUSTER", 3)

    h = df4["high"].to_numpy(dtype=float)
    l = df4["low"].to_numpy(dtype=float)
    c = df4["close"].to_numpy(dtype=float)
    o = df4["open"].to_numpy(dtype=float)
    v = df4["volume"].to_numpy(dtype=float)

    n = len(c)
    if n < 220:
        # need enough bars for stability (BB/MFI/WT/ATR)
        return None

    # --- WT ---
    src_wt = (h + l + c) / 3.0
    wt_esa = _ema(src_wt, wt_n1)
    wt_d = _ema(np.abs(src_wt - wt_esa), wt_n1)
    wt_ci = (src_wt - wt_esa) / (0.015 * np.where(wt_d == 0, np.nan, wt_d))
    wt_tci = _ema(wt_ci, wt_n2)
    wt_sig = _sma(wt_tci, wt_n3)

    # --- MFI ---
    mfi_raw = _mfi(h, l, c, v, mfi_len)
    mfi_val = mfi_raw if mfi_smooth <= 1 else _sma(mfi_raw, mfi_smooth)

    # --- BB ---
    bb_basis = _sma(c, bb_len)
    bb_dev = _stdev(c, bb_len) * bb_mult
    bb_up = bb_basis + bb_dev
    bb_dn = bb_basis - bb_dev

    # --- SuperTrend (state + hysteresis + confirm) ---
    atr = _atr(h, l, c, st_atr_len)
    st_mid = (h + l) * 0.5
    upperBasic = st_mid + st_factor * atr
    lowerBasic = st_mid - st_factor * atr

    st_upper = np.nan
    st_lower = np.nan
    st_state = None  # 1 or -1

    st_candidate = None
    cand_count = 0
    lock_band = np.nan
    lock_h = np.nan

    # --- State vars for arms/signals ---
    wt_arm_dn = False
    wt_arm_up = False
    sig_wt_buy = False

    mfi_armL = False
    mfi_armH = False
    mfi_armL_age = 0
    mfi_armH_age = 0
    mfi_allow_buy = True
    mfi_allow_sell = True
    sig_mfi_buy = False

    bb_buy_armed = False
    bb_buy_in_cnt = 0
    bb_buy_age = 0
    sig_bb_buy = False

    last_wt_buy = None
    last_mfi_buy = None
    last_bb_buy = None

    # iterate bars (confirmed = every historical closed bar)
    for i in range(n):
        # skip until we have required indicator values
        if np.isnan(wt_tci[i]) or np.isnan(wt_sig[i]) or np.isnan(mfi_val[i]) or np.isnan(bb_dn[i]) or np.isnan(bb_up[i]) or np.isnan(atr[i]):
            continue

        confirmed = True

        # ---- WT engine parity ----
        sig_wt_buy = False
        wt_touch_top = wt_tci[i] >= wt_ob
        wt_touch_bot = wt_tci[i] <= wt_os
        wt_cross_up = (wt_tci[i-1] <= wt_sig[i-1]) and (wt_tci[i] > wt_sig[i]) if i > 0 and not np.isnan(wt_tci[i-1]) and not np.isnan(wt_sig[i-1]) else False
        wt_cross_dn = (wt_tci[i-1] >= wt_sig[i-1]) and (wt_tci[i] < wt_sig[i]) if i > 0 and not np.isnan(wt_tci[i-1]) and not np.isnan(wt_sig[i-1]) else False

        if confirmed:
            if wt_touch_top:
                wt_arm_up = True
                wt_arm_dn = False
            if wt_touch_bot:
                wt_arm_dn = True
                wt_arm_up = False

            if wt_arm_dn and wt_cross_up:
                sig_wt_buy = True
                wt_arm_dn = False

            if wt_arm_up and wt_cross_dn:
                wt_arm_up = False

        if sig_wt_buy:
            last_wt_buy = i

        # ---- MFI engine parity ----
        sig_mfi_buy = False
        mfi_zone_low = mfi_val[i] <= mfi_os
        mfi_zone_high = mfi_val[i] >= mfi_ob

        pivot_up = False
        pivot_down = False
        if i >= 2 and not np.isnan(mfi_val[i-2]) and not np.isnan(mfi_val[i-1]):
            d_prev = mfi_val[i-1] - mfi_val[i-2]
            d_now = mfi_val[i] - mfi_val[i-1]
            pivot_up = (d_prev < 0) and (d_now > 0)
            pivot_down = (d_prev > 0) and (d_now < 0)

        if confirmed:
            if mfi_zone_low:
                mfi_armL = True
                mfi_armH = False
                mfi_armL_age = 0
                mfi_armH_age = 0
                mfi_allow_buy = True
                mfi_allow_sell = True
            elif mfi_armL and (not mfi_zone_low):
                mfi_armL_age += 1
                if mfi_armL_age > mfi_linger:
                    mfi_armL = False
                    mfi_armL_age = 0

            if mfi_zone_high:
                mfi_armH = True
                mfi_armL = False
                mfi_armH_age = 0
                mfi_armL_age = 0
                mfi_allow_buy = True
                mfi_allow_sell = True
            elif mfi_armH and (not mfi_zone_high):
                mfi_armH_age += 1
                if mfi_armH_age > mfi_linger:
                    mfi_armH = False
                    mfi_armH_age = 0

            if pivot_up:
                mfi_allow_sell = True
            if pivot_down:
                mfi_allow_buy = True

            if mfi_armL and mfi_allow_buy and pivot_up:
                sig_mfi_buy = True
                mfi_allow_buy = False
                mfi_armL = False
                mfi_armL_age = 0

            if mfi_armH and mfi_allow_sell and pivot_down:
                mfi_allow_sell = False
                mfi_armH = False
                mfi_armH_age = 0

        if sig_mfi_buy:
            last_mfi_buy = i

        # ---- BB engine parity ----
        sig_bb_buy = False
        if confirmed:
            touch_dn = (l[i] <= bb_dn[i]) and (bb_dn[i] <= h[i])
            outside_dn = c[i] < bb_dn[i]
            inside_close = (c[i] >= bb_dn[i]) and (c[i] <= bb_up[i])

            if (not bb_buy_armed) and (touch_dn or outside_dn):
                bb_buy_armed = True
                bb_buy_in_cnt = 0
                bb_buy_age = 0

            if bb_buy_armed:
                bb_buy_age += 1
                if inside_close:
                    bb_buy_in_cnt += 1
                    if bb_buy_in_cnt >= bb_rein:
                        sig_bb_buy = True
                        bb_buy_armed = False
                        bb_buy_in_cnt = 0
                        bb_buy_age = 0
                else:
                    bb_buy_in_cnt = 0

                if bb_buy_armed and (bb_buy_age >= bb_reentry):
                    bb_buy_armed = False
                    bb_buy_in_cnt = 0
                    bb_buy_age = 0

        if sig_bb_buy:
            last_bb_buy = i

        # ---- SuperTrend state machine (Pine parity-ish) ----
        if st_state is None:
            st_state = 1 if c[i] >= st_mid[i] else -1
            st_upper = upperBasic[i]
            st_lower = lowerBasic[i]
            st_candidate = None
            cand_count = 0
            lock_band = np.nan
            lock_h = np.nan
        else:
            # band update (trail only on active trend)
            if st_state == 1:
                st_lower = lowerBasic[i] if np.isnan(st_lower) else max(lowerBasic[i], st_lower)
                st_upper = upperBasic[i]
            else:
                st_upper = upperBasic[i] if np.isnan(st_upper) else min(upperBasic[i], st_upper)
                st_lower = lowerBasic[i]

            # flip logic
            if confirmed and i > 0:
                ref_line = st_lower if st_state == 1 else st_upper
                atr_prev = atr[i-1] if i > 0 and not np.isnan(atr[i-1]) else atr[i]
                st_hys_dist = st_hys_atr * atr_prev

                start_dn = (st_state == 1) and (not np.isnan(ref_line)) and (c[i] < (ref_line - st_hys_dist))
                start_up = (st_state == -1) and (not np.isnan(ref_line)) and (c[i] > (ref_line + st_hys_dist))

                if st_candidate is None:
                    if start_dn:
                        st_candidate = -1
                        lock_band = ref_line
                        lock_h = st_hys_dist
                        cand_count = 1
                    elif start_up:
                        st_candidate = 1
                        lock_band = ref_line
                        lock_h = st_hys_dist
                        cand_count = 1
                else:
                    pass_cond = (c[i] < (lock_band - lock_h)) if st_candidate == -1 else (c[i] > (lock_band + lock_h)) if st_candidate == 1 else False
                    if pass_cond:
                        cand_count += 1
                        if cand_count >= st_confirm:
                            st_state = st_candidate
                            st_candidate = None
                            cand_count = 0
                            lock_band = np.nan
                            lock_h = np.nan
                            # reset bands on flip
                            st_upper = upperBasic[i]
                            st_lower = lowerBasic[i]
                    else:
                        st_candidate = None
                        cand_count = 0
                        lock_band = np.nan
                        lock_h = np.nan

    # Snapshot at the last bar with valid values
    i = n - 1
    # Find last index where indicators exist
    while i > 0 and (np.isnan(wt_tci[i]) or np.isnan(wt_sig[i]) or np.isnan(mfi_val[i]) or np.isnan(bb_dn[i]) or np.isnan(bb_up[i]) or np.isnan(atr[i])):
        i -= 1
    if i <= 0 or st_state is None:
        return None

    st_dn = (st_state == -1)
    arm_flags = {
        "WT": bool(wt_arm_dn),
        "MFI": bool(mfi_armL),
        "BB": bool(bb_buy_armed),
    }
    arm_cnt = int(sum(1 for k, v in arm_flags.items() if v))

    # cluster count (info only)
    def within(last_idx: Optional[int]) -> bool:
        return (last_idx is not None) and (i - last_idx <= ttl)

    buy_cnt = (1 if within(last_wt_buy) else 0) + (1 if within(last_mfi_buy) else 0) + (1 if within(last_bb_buy) else 0)

    # simple ATR% for risk
    atrpct = float((atr[i] / c[i]) * 100.0) if c[i] > 0 else np.nan

    return {
        "st_dn": st_dn,
        "arm_flags": arm_flags,
        "arm_cnt": arm_cnt,
        "buy_cnt": buy_cnt,
        "atrpct_4h": atrpct,
        "wt_tci": float(wt_tci[i]),
        "mfi": float(mfi_val[i]),
        "bb_pos": "OUT" if c[i] < bb_dn[i] else "IN" if (c[i] >= bb_dn[i] and c[i] <= bb_up[i]) else "UP",
        "bars_4h": int(len(df4)),
    }


# -----------------------
# S3: Candidate selection (daily prefilter + 4H main setup)
# -----------------------
def daily_prefilter(ticker: str) -> Optional[dict]:
    """
    Cheap daily filter (Stooq): drawdown + (1W+) + (1M-) + liquidity.
    This is not the final selection, only for reducing yfinance calls.
    """
    timeout = env_int("DATA_TIMEOUT", 12)
    series = _fetch_stooq_daily_ohlcv(ticker, timeout=timeout)
    if not series or len(series) < 260:
        return None
    closes = _closes(series)
    close = closes[-1]
    if close <= 0:
        return None

    # Liquidity (best-effort: Stooq volume)
    vol20 = sum([x["volume"] for x in series[-20:]]) / 20.0
    dv20 = vol20 * close
    if close < env_float("MIN_PRICE", 5):
        return None
    if dv20 < env_float("MIN_DOLLAR_VOL", 15000000):
        return None

    high252 = max(closes[-252:])
    dd = (close / high252) - 1.0  # negative

    dd_soft = env_float("DD_SOFT", -0.15)
    if dd > dd_soft:
        return None

    r5 = _pct_change(closes, 5)
    r21 = _pct_change(closes, 21)

    if env("REQUIRE_1W_POS", "1") != "0":
        if r5 is None or r5 <= 0:
            return None
    if env("REQUIRE_1M_NEG", "1") != "0":
        if r21 is None or r21 >= 0:
            return None

    return {"dd": dd, "r5": r5, "r21": r21, "dv20": dv20}


def build_s3_with_main_indicator(chosen_etfs: List[str], state: dict) -> Tuple[List[str], List[str], List[dict]]:
    warnings: List[str] = []
    timeout = env_int("DATA_TIMEOUT", 12)

    holdings_per_etf = env_int("S3_HOLDINGS_PER_ETF", 40)
    max_universe = env_int("S3_MAX_UNIVERSE", 160)

    top_n = env_int("S3_TOP_N", 4)
    base_top_n = env_int("S3_BASE_TOP_N", 4)

    cooldown_days = env_int("S3_COOLDOWN_DAYS", 7)

    require_st_down = env("REQUIRE_ST_DOWN", "1") != "0"
    a_min_arm = env_int("A_MIN_ARM", 2)
    b_min_arm = env_int("B_MIN_ARM", 1)

    dd_hard = env_float("DD_HARD", -0.25)
    dd_soft = env_float("DD_SOFT", -0.15)

    # Debug counters
    dbg = {
        "etf": 0, "hold": 0, "universe": 0,
        "prefilter_ok": 0, "yf_ok": 0,
        "st_dn": 0, "A": 0, "B": 0,
        "cooldown_skip": 0
    }

    # Build universe from holdings
    universe: Dict[str, dict] = {}
    for etf in chosen_etfs:
        try:
            hs = fetch_spdr_holdings(etf, timeout=timeout)
            dbg["etf"] += 1
            if not hs:
                warnings.append(f"S3 holdings missing: {etf}")
                continue
            dbg["hold"] += min(len(hs), holdings_per_etf)
            for h in hs[:holdings_per_etf]:
                t = h["ticker"]
                if t not in universe:
                    universe[t] = {"name": h.get("name", ""), "from": set([etf])}
                else:
                    universe[t]["from"].add(etf)
        except Exception:
            warnings.append(f"S3 holdings error: {etf}")
            print(traceback.format_exc(), file=sys.stderr)

    tickers = list(universe.keys())[:max_universe]
    dbg["universe"] = len(tickers)
    if not tickers:
        return [tr("s3_none"), tr("s3_note")], warnings, []

    # Step 1: daily prefilter to reduce yfinance calls
    pre = []
    for t in tickers:
        try:
            m = daily_prefilter(t)
            if not m:
                continue
            m["ticker"] = t
            m["name"] = universe[t].get("name", "")
            m["from"] = ",".join(sorted(list(universe[t]["from"])))
            pre.append(m)
        except Exception:
            continue

    dbg["prefilter_ok"] = len(pre)
    # prioritize deeper drawdown first
    pre.sort(key=lambda x: x["dd"])
    pre = pre[: min(len(pre), 40)]  # cap yfinance calls

    yf_period = env("YF_PERIOD", "180d") or "180d"
    yf_interval = env("YF_INTERVAL", "1h") or "1h"

    A_list: List[dict] = []
    B_list: List[dict] = []

    for it in pre:
        t = it["ticker"]

        # cooldown (skip except if it's really strong later)
        if was_recently_recommended(state, t, cooldown_days):
            dbg["cooldown_skip"] += 1
            continue

        df1 = yf_fetch_1h(t, period=yf_period, interval=yf_interval)
        if df1 is None:
            continue
        df4 = resample_4h(df1)
        if df4 is None:
            continue

        setup = compute_main_setup(df4)
        if setup is None:
            continue

        dbg["yf_ok"] += 1
        if setup["st_dn"]:
            dbg["st_dn"] += 1

        if require_st_down and (not setup["st_dn"]):
            continue

        arm_cnt = setup["arm_cnt"]
        dd = it["dd"]
        # classify A/B
        is_A = (arm_cnt >= a_min_arm) and (dd <= dd_hard)
        is_B = (arm_cnt >= b_min_arm) and (dd <= dd_soft) and (not is_A)

        # If too strict, allow A by arm only (keep it meaningful)
        if (not is_A) and (arm_cnt >= a_min_arm) and (dd <= dd_soft) and (setup["buy_cnt"] >= 2):
            is_A = True
            is_B = False

        # scoring: arm + buy cluster + drawdown depth - volatility penalty
        atrp = setup.get("atrpct_4h", np.nan)
        vol_pen = 0.0
        if not np.isnan(atrp):
            # sweet spot around 3~6%
            vol_pen = abs(atrp - 4.5) / 6.0  # 0..~1
            vol_pen = min(max(vol_pen, 0.0), 1.0)

        dd_score = min(max((-dd) / 0.6, 0.0), 1.0)
        score = 100.0 * (0.45 * (arm_cnt / 3.0) + 0.25 * (setup["buy_cnt"] / 3.0) + 0.25 * dd_score + 0.05 * (1.0 - vol_pen))
        grade = "SSS" if score >= 90 else "SS" if score >= 80 else "S" if score >= 70 else "C"

        rec = {
            "ticker": t,
            "name": it.get("name", ""),
            "from": it.get("from", ""),
            "dd": dd,
            "r5": it.get("r5", None),
            "r21": it.get("r21", None),
            "arm_cnt": arm_cnt,
            "arms": setup["arm_flags"],
            "buy_cnt": setup["buy_cnt"],
            "atrp4h": atrp,
            "score": score,
            "grade": grade,
        }

        if is_A:
            A_list.append(rec)
            dbg["A"] += 1
        elif is_B:
            B_list.append(rec)
            dbg["B"] += 1

    A_list.sort(key=lambda x: x["score"], reverse=True)
    B_list.sort(key=lambda x: x["score"], reverse=True)

    A_pick = A_list[:top_n]
    B_pick = B_list[:base_top_n]

    lines: List[str] = []
    today_str = now_kst().strftime("%Y-%m-%d")
    hist_items: List[dict] = []

    def fmt_pct(x: Optional[float]) -> str:
        if x is None:
            return "NA"
        sign = "+" if x >= 0 else ""
        return f"{sign}{x:.1f}%"

    def arms_str(arms: Dict[str, bool]) -> str:
        on = [k for k, v in arms.items() if v]
        return "+".join(on) if on else "-"

    lines.append(f"<b>{html.escape(tr('s3a'))}</b>")
    if not A_pick:
        lines.append("• (없음)")
    else:
        for i, r in enumerate(A_pick, 1):
            t = r["ticker"]
            name = (r.get("name","") or "").strip()
            name_short = (name[:28] + "…") if len(name) > 28 else name
            ddp = r["dd"] * 100.0
            lines.append(f"{i}) <b>[{r['grade']}] {html.escape(t)}</b> {html.escape(name_short)}")
            lines.append(f"   • ST:DOWN / arm({r['arm_cnt']}): {html.escape(arms_str(r['arms']))} / cluster(3bars): {r['buy_cnt']}/3")
            lines.append(f"   • 급락(52W): {ddp:.1f}% / 1W: {fmt_pct(r.get('r5'))} / 1M: {fmt_pct(r.get('r21'))}")
            lines.append(f"   • 리스크(4H): ATR% {r.get('atrp4h', float('nan')):.1f} / 출처(ETF): {html.escape(r.get('from',''))} / 점수: {r['score']:.0f}")
            hist_items.append({"date_kst": today_str, "ticker": t, "grade": r["grade"], "score": int(r["score"])})

    lines.append("")
    lines.append(f"<b>{html.escape(tr('s3b'))}</b>")
    if not B_pick:
        lines.append("• (없음)")
    else:
        for i, r in enumerate(B_pick, 1):
            t = r["ticker"]
            name = (r.get("name","") or "").strip()
            name_short = (name[:28] + "…") if len(name) > 28 else name
            ddp = r["dd"] * 100.0
            lines.append(f"{i}) <b>[{r['grade']}] {html.escape(t)}</b> {html.escape(name_short)}")
            lines.append(f"   • ST:DOWN / arm({r['arm_cnt']}): {html.escape(arms_str(r['arms']))} / cluster(3bars): {r['buy_cnt']}/3")
            lines.append(f"   • 급락(52W): {ddp:.1f}% / 1W: {fmt_pct(r.get('r5'))} / 1M: {fmt_pct(r.get('r21'))}")
            lines.append(f"   • 리스크(4H): ATR% {r.get('atrp4h', float('nan')):.1f} / 출처(ETF): {html.escape(r.get('from',''))} / 점수: {r['score']:.0f}")

    lines.append(tr("s3_note"))

    if hist_items:
        add_candidate_history(state, hist_items)

    if env("DEBUG_S3", "0") != "0":
        warnings.append(
            "S3 스캔: "
            f"ETF={dbg['etf']}, holdings={dbg['hold']}, universe={dbg['universe']}, "
            f"prefilter={dbg['prefilter_ok']}, yfinance_ok={dbg['yf_ok']}, st_down={dbg['st_dn']}, "
            f"A={dbg['A']}, B={dbg['B']}, cooldown_skip={dbg['cooldown_skip']}"
        )

    return lines, warnings, hist_items


# -----------------------
# Telegram
# -----------------------
def send_telegram(text: str) -> None:
    token = env("TELEGRAM_BOT_TOKEN")
    chat_id = env("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("Missing TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID")

    url = f"{TELEGRAM_API}/bot{token}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML", "disable_web_page_preview": True}
    resp = requests.post(url, json=payload, timeout=12)
    if resp.status_code != 200:
        raise RuntimeError(f"Telegram send failed: {resp.status_code} {resp.text}")


# -----------------------
# Main build message
# -----------------------
def build_message(etf_map: dict, state: dict) -> str:
    logs: List[str] = []
    kst = now_kst().strftime("%Y-%m-%d %H:%M:%S KST")
    sha = env("GITHUB_SHA")
    sha_short = sha[:7] if sha else "local"

    benchmark = (etf_map.get("benchmark", []) or ["ACWI"])[0]
    sectors = etf_map.get("sectors_11", []) or []
    cr = etf_map.get("countries_regions_16", {}) or {}
    countries_cnt = len((cr.get("developed", []) or [])) + len((cr.get("emerging", []) or []))

    lines: List[str] = []
    lines.append(f"<b>{html.escape(tr('title'))}</b>  <code>{html.escape(tr('status_ok'))}</code>")
    lines.append(f"⏱️ {html.escape(kst)}")
    lines.append(f"🧩 v{VERSION} / {html.escape(sha_short)}")
    lines.append(f"🗺️ {html.escape(tr('etf_loaded'))}: benchmark={html.escape(benchmark)} / sectors={len(sectors)} / countries={countries_cnt}")
    lines.append("")

    # S2
    sector_a: List[str] = []
    sector_b: List[str] = []
    country_a: List[str] = []
    country_b: List[str] = []
    try:
        sector_a, sector_b, country_a, country_b, w = compute_s2_picks(etf_map)
        logs.extend(w)
        lines.extend(build_s2_section(etf_map, sector_a, sector_b, country_a, country_b))
        lines.append("")
    except Exception:
        logs.append("S2 failed")
        print(traceback.format_exc(), file=sys.stderr)

    # S3 (main-indicator scan)
    chosen: List[str] = []
    if sector_a:
        chosen.append(sector_a[0])
    if sector_b and sector_b[0] not in chosen:
        chosen.append(sector_b[0])

    try:
        s3_lines, w, _ = build_s3_with_main_indicator(chosen, state)
        logs.extend(w)
        lines.extend(s3_lines)
        lines.append("")
    except Exception:
        lines.append(f"<b>{html.escape(tr('s3a'))}</b>")
        lines.append(html.escape(tr("s3_none")))
        lines.append("")
        logs.append("S3 failed")
        print(traceback.format_exc(), file=sys.stderr)

    # S1
    rss_urls = load_rss_urls()
    headlines, rss_logs = fetch_headlines(rss_urls, top_n=16, timeout=12, state=state)
    logs.extend(rss_logs)
    digest_lines, used_links = build_news_digest(headlines)
    lines.extend(digest_lines)
    if used_links:
        mark_links_seen(state, used_links)

    state["last_run_kst"] = now_kst().strftime("%Y-%m-%d %H:%M:%S")
    prune_state(state)

    if logs:
        lines.append("")
        lines.append(f"<b>{html.escape(tr('s4_log'))}</b>")
        for w in logs[:10]:
            lines.append(f"• {html.escape(w)}")

    return truncate_telegram("\n".join(lines))


def main() -> int:
    state, _ = load_state()
    etf_map = load_etf_map()

    try:
        msg = build_message(etf_map, state)
        save_state(state)
        send_telegram(msg)
        print("[OK] Briefing sent.")
        return 0
    except Exception as e:
        err = f"{type(e).__name__}: {str(e)[:300]}"
        print("[ERROR] " + err, file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        try:
            kst = now_kst().strftime("%Y-%m-%d %H:%M:%S KST")
            text = f"<b>{html.escape(tr('title'))}</b>  <code>{html.escape(tr('status_err'))}</code>\n⏱️ {html.escape(kst)}\n<b>{html.escape(tr('s4_err'))}</b>\n<code>{html.escape(err)}</code>"
            send_telegram(text)
        except Exception:
            pass
        return 1


if __name__ == "__main__":
    raise SystemExit(main())

